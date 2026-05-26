#!/usr/bin/env python3
"""
layers_usuarios.py — Extrai todos os usuários (ativos e inativos) de uma comunidade Layers

CAMPOS RETORNADOS PELA API:
  status       = "ACTIVE"  → usuário ativou e usou o app pelo menos 1x
               = "INVITED" → usuário recebeu convite mas nunca ativou a conta
  activatedAt  = data de ativação (null = nunca ativou)
  lastSeenAt   = último acesso (null = nunca acessou)

AUTENTICAÇÃO:
  O token do tipo "auth:app" só funciona nas comunidades onde o app está instalado.
  - Para usar outra comunidade: gerar token admin no painel da Layers
  - Ou instalar o app na comunidade desejada

USO:
  # Uma comunidade, exportar para XLSX
  python layers_usuarios.py --community americano --token SEU_TOKEN

  # Comunidade padrão (raizeducacao), variável de ambiente
  export LAYERS_API_TOKEN=eyJ...
  python layers_usuarios.py --community raizeducacao

  # Múltiplas comunidades
  python layers_usuarios.py --community americano qi-tijuca leonardodavinci-alfa --token SEU_TOKEN

  # Saída em JSON
  python layers_usuarios.py --community raizeducacao --format json

  # Só inativos
  python layers_usuarios.py --community raizeducacao --status INVITED

DEPENDÊNCIAS:
  pip install requests pandas openpyxl
"""

import os
import json
import time
import argparse
from datetime import datetime
from pathlib import Path

try:
    import requests
except ImportError:
    print("Instale requests: pip install requests")
    raise

try:
    import pandas as pd
except ImportError:
    print("Instale pandas: pip install pandas openpyxl")
    raise

# ─── Configuração ─────────────────────────────────────────────────────────────

LAYERS_BASE_URL = "https://api.layers.digital"
PAGE_SIZE = 200          # usuários por request (max ~1000, 200 é seguro)
RATE_LIMIT_DELAY = 0.2  # segundos entre requests

# Mapeamento community_id → nome amigável
COMMUNITY_NAMES = {
    "raizeducacao":         "Raiz Educação (Holding)",
    "americano":            "Colégio Americano Bilíngue",
    "leonardodavinci-alfa": "Leonardo da Vinci - Alfa",
    "leonardodavinci-beta": "Leonardo da Vinci - Beta",
    "leonardodavinci-gama": "Leonardo da Vinci - Gama",
    "qi-tijuca":            "QI Tijuca",
    "qi-recreio":           "QI Recreio",
    "qi-rio2":              "QI Rio 2",
    "qi-metropolitano":     "QI Metropolitano",
    "qi-freguesia":         "QI Freguesia",
    "az51800x":             "QI Valqueire",
    "matriz-bangu":         "Matriz Bangu",
    "matriz-campogrande":   "Matriz Campo Grande",
    "matriz-caxias":        "Matriz Caxias",
    "matriz-madureira":     "Matriz Madureira",
    "matriz-novaiguacu":    "Matriz Nova Iguaçu",
    "matriz-tijuca":        "Matriz Tijuca",
    "uniao":                "Colégio União",
    "sap":                  "Escola SAP",
    "xa7y5zam":             "Sá Pereira - Fund/Médio",
    "w213sfza":             "Sá Pereira - Infantil",
    "bomtempo":             "Bom Tempo",
    "sarahdawsey-juizdefora": "Sarah Dawsey - JF",
    "y9490m37":             "Sara Dawsey - Tijuca",
}

# ─── Funções de API ───────────────────────────────────────────────────────────

def fetch_page(token: str, community_id: str, skip: int, limit: int) -> list[dict]:
    url = f"{LAYERS_BASE_URL}/v1/users"
    headers = {
        "Authorization": f"Bearer {token}",
        "community-id":  community_id,
    }
    resp = requests.get(
        url,
        headers=headers,
        params={"limit": limit, "skip": skip},
        timeout=30,
    )

    if resp.status_code == 401:
        raise PermissionError(
            f"Sem acesso à comunidade '{community_id}'. "
            "Verifique se o app está instalado lá ou use um token admin."
        )
    if resp.status_code == 429:
        print("  Rate limit — aguardando 10s...")
        time.sleep(10)
        return fetch_page(token, community_id, skip, limit)

    resp.raise_for_status()
    data = resp.json()
    return data if isinstance(data, list) else []


def fetch_all_users(token: str, community_id: str) -> list[dict]:
    all_users: list[dict] = []
    skip = 0

    print(f"  [{community_id}] buscando...", end="", flush=True)
    while True:
        page = fetch_page(token, community_id, skip, PAGE_SIZE)
        all_users.extend(page)
        print(f" {len(all_users)}", end="", flush=True)

        if len(page) < PAGE_SIZE:
            break
        skip += PAGE_SIZE
        time.sleep(RATE_LIMIT_DELAY)

    print()
    return all_users


def normalize(user: dict, community_id: str) -> dict:
    roles = user.get("roles") or []
    tipo = "admin" if "admin" in roles else (
        "aluno"       if "student" in roles else (
        "responsavel" if "guardian" in roles else
        ", ".join(roles) or "—"
    ))

    return {
        "comunidade":          COMMUNITY_NAMES.get(community_id, community_id),
        "community_id":        community_id,
        "layers_id":           user.get("_id", ""),
        "nome":                user.get("name", ""),
        "email":               user.get("email", ""),
        "status":              user.get("status", ""),
        "ativo":               "Sim" if user.get("status") == "ACTIVE" else "Não",
        "tipo":                tipo,
        "criado_em":           _fmt_date(user.get("createdAt")),
        "ativado_em":          _fmt_date(user.get("activatedAt")),
        "ultimo_acesso":       _fmt_date(user.get("lastSeenAt")),
        "qtd_convites":        user.get("invitationCount", 0),
        "ultimo_convite":      _fmt_date(user.get("lastSentInvitation")),
    }


def _fmt_date(val: str | None) -> str:
    if not val:
        return ""
    try:
        dt = datetime.fromisoformat(val.replace("Z", "+00:00"))
        return dt.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return val


# ─── Exportação ───────────────────────────────────────────────────────────────

def export_xlsx(records: list[dict], path: str) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    df = pd.DataFrame(records)
    df.to_excel(path, index=False)

    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Cabeçalho em negrito com fundo azul
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Colorir linhas de inativos em amarelo claro
    yellow = PatternFill("solid", fgColor="FFF2CC")
    for row in ws.iter_rows(min_row=2):
        status_cell = row[5]  # coluna "status"
        if status_cell.value == "INVITED":
            for cell in row:
                cell.fill = yellow

    # Ajustar largura das colunas
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"
    wb.save(path)


# ─── Main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Extrai usuários ativos/inativos da Layers API",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "--community", "-c",
        nargs="+",
        required=True,
        help="ID(s) da(s) comunidade(s). Ex: raizeducacao americano",
    )
    parser.add_argument(
        "--token", "-t",
        default=os.getenv("LAYERS_API_TOKEN", ""),
        help="Bearer token da Layers API (ou LAYERS_API_TOKEN env var)",
    )
    parser.add_argument(
        "--format", "-f",
        choices=["xlsx", "json", "csv"],
        default="xlsx",
        help="Formato de saída (default: xlsx)",
    )
    parser.add_argument(
        "--output", "-o",
        default="",
        help="Nome do arquivo de saída (default: auto gerado)",
    )
    parser.add_argument(
        "--status",
        choices=["ACTIVE", "INVITED", "all"],
        default="all",
        help="Filtrar por status (default: all)",
    )
    args = parser.parse_args()

    if not args.token:
        print("ERRO: token não encontrado. Use --token ou defina LAYERS_API_TOKEN.")
        return

    all_records: list[dict] = []

    for community_id in args.community:
        try:
            users = fetch_all_users(args.token, community_id)
            records = [normalize(u, community_id) for u in users]

            if args.status != "all":
                records = [r for r in records if r["status"] == args.status]

            all_records.extend(records)

            # Sumário por comunidade
            ativos    = sum(1 for r in records if r["status"] == "ACTIVE")
            inativos  = len(records) - ativos
            print(f"  >> {len(records)} usuarios: {ativos} ativos, {inativos} inativos/convidados")

        except PermissionError as e:
            print(f"  ACESSO NEGADO: {e}")
        except Exception as e:
            print(f"  ERRO em '{community_id}': {e}")

    if not all_records:
        print("Nenhum dado para exportar.")
        return

    # Sumário final
    print(f"\nTOTAL: {len(all_records)} usuarios em {len(args.community)} comunidade(s)")
    ativos_total = sum(1 for r in all_records if r["status"] == "ACTIVE")
    print(f"  Ativos:   {ativos_total}")
    print(f"  Inativos: {len(all_records) - ativos_total}")

    # Nome do arquivo de saída
    suffix = datetime.now().strftime("%Y%m%d_%H%M%S")
    communities_slug = "_".join(args.community[:2])
    output = args.output or f"layers_{communities_slug}_{suffix}.{args.format}"

    # Exportar
    if args.format == "json":
        with open(output, "w", encoding="utf-8") as f:
            json.dump(all_records, f, ensure_ascii=False, indent=2)
    elif args.format == "csv":
        df = pd.DataFrame(all_records)
        df.to_csv(output, index=False, encoding="utf-8-sig")
    else:
        export_xlsx(all_records, output)

    print(f"\nExportado: {output}")


if __name__ == "__main__":
    main()
