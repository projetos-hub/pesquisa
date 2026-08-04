#!/usr/bin/env python3
'''Read-only safety CLI for Pesquisa Miniapp Layers.'''
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import sys
import unicodedata
import urllib.error
import urllib.request
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

APP_URL = 'https://pesquisa-nu-sand.vercel.app'
PROJECT_ID = 'qnpvlhfjknnvfiyxrhhl'
SKILL_NAME = 'pesquisa-miniapp-layers'

CAPABILITIES = {
    'survey': ['list', 'show', 'create', 'update', 'duplicate', 'delete', 'status'],
    'question': ['list', 'create', 'update', 'move', 'replace-options', 'delete', 'flow'],
    'community': ['list', 'identity', 'install', 'update-installation', 'remove', 'dates', 'theme', 'text-override'],
    'sample': ['audit', 'prepare', 'append', 'replace', 'clear', 'resolve', 'group-create', 'group-members', 'group-delete'],
    'template': ['list', 'create', 'update', 'clone', 'validate', 'archive'],
    'dispatch': ['preview', 'create', 'schedule', 'process', 'status', 'retry', 'cancel'],
    'comunicado': ['list', 'create', 'update', 'publish', 'archive'],
    'link': ['list', 'create', 'scope', 'disable', 'rotate-key'],
    'response': ['count', 'list', 'export', 'audit'],
    'report': ['overview', 'questions', 'segments', 'communities', 'timeline', 'funnel', 'xlsx'],
    'platform': ['doctor', 'health', 'quality', 'migration', 'rls', 'deploy', 'smoke'],
}

DESTRUCTIVE = {'survey.delete', 'community.remove', 'sample.replace', 'sample.clear', 'sample.group-delete', 'link.disable'}
EXTERNAL = {'dispatch.create', 'dispatch.schedule', 'dispatch.process', 'dispatch.retry', 'comunicado.publish'}
SCHEMA_OR_DEPLOY = {'platform.deploy', 'platform.migration', 'platform.rls'}
FORBIDDEN_KEY_PARTS = ('password', 'secret', 'service_role', 'token', 'access_key', 'authorization')
MOJIBAKE_MARKERS = ('\ufffd', '\x00', '\u00c3', '\u00c2', '\u00e2\u20ac', '\u00ef\u00bf\u00bd')
EMAIL_RE = re.compile(r'^[^\s@]+@[^\s@]+\.[^\s@]+$')
SUSPICIOUS_Q_RE = re.compile(r'[A-Za-z\u00c0-\u00ff]\?[A-Za-z\u00c0-\u00ff]|\?{2,}')


def skill_root() -> Path:
    return Path(__file__).resolve().parents[1]


def json_print(value: Any) -> None:
    print(json.dumps(value, ensure_ascii=False, indent=2, default=str))


def env_presence(path: Path) -> dict[str, bool]:
    names = {
        'NEXT_PUBLIC_SUPABASE_URL', 'NEXT_PUBLIC_SUPABASE_PUBLISHABLE_KEY',
        'SUPABASE_SERVICE_ROLE_KEY', 'LAYERS_API_TOKEN', 'CRON_SECRET',
        'NEXT_PUBLIC_APP_URL', 'SHEETS_WEBHOOK_URL', 'SHEETS_WEBHOOK_SECRET',
    }
    found: set[str] = set()
    if path.exists():
        for raw in path.read_text(encoding='utf-8-sig').splitlines():
            line = raw.strip()
            if line and not line.startswith('#') and '=' in line:
                found.add(line.split('=', 1)[0].strip())
    found.update(name for name in names if os.environ.get(name))
    return {name: name in found for name in sorted(names)}


def text_issues(text: str, source: str = 'text') -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    for marker in MOJIBAKE_MARKERS:
        if marker in text:
            issues.append({'source': source, 'kind': 'mojibake', 'marker': marker.encode('unicode_escape').decode()})
    if unicodedata.normalize('NFC', text) != text:
        issues.append({'source': source, 'kind': 'not_nfc'})
    if SUSPICIOUS_Q_RE.search(text):
        issues.append({'source': source, 'kind': 'suspicious_question_mark'})
    controls = [ord(ch) for ch in text if ord(ch) < 32 and ch not in '\t\r\n']
    if controls:
        issues.append({'source': source, 'kind': 'control_char', 'codes': sorted(set(controls))})
    return issues


def walk_strings(value: Any, path: str = '$') -> Iterable[tuple[str, str]]:
    if isinstance(value, str):
        yield path, value
    elif isinstance(value, dict):
        for key, item in value.items():
            yield from walk_strings(item, f'{path}.{key}')
    elif isinstance(value, list):
        for index, item in enumerate(value):
            yield from walk_strings(item, f'{path}[{index}]')


def iter_file_strings(path: Path, sheet: str | None = None) -> Iterable[tuple[str, str]]:
    suffix = path.suffix.lower()
    if suffix in {'.txt', '.md', '.html', '.csv', '.json', '.py', '.yaml', '.yml', '.ts', '.tsx'}:
        text = path.read_text(encoding='utf-8-sig', errors='strict')
        if suffix == '.json':
            yield from walk_strings(json.loads(text), str(path))
        else:
            for number, line in enumerate(text.splitlines(), 1):
                yield f'{path}:{number}', line
        return
    if suffix == '.xlsx':
        try:
            import openpyxl
        except ImportError as exc:
            raise SystemExit('Instale openpyxl para auditar XLSX.') from exc
        book = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = book[sheet] if sheet else book.active
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    yield f'{path}:{ws.title}!{cell.coordinate}', cell.value
        book.close()
        return
    raise SystemExit(f'Formato nao suportado: {suffix}')


def command_doctor(args: argparse.Namespace) -> int:
    root = skill_root()
    required = [
        root / 'SKILL.md',
        root / 'agents' / 'openai.yaml',
        root / 'scripts' / 'external_client.py',
        root / 'references' / 'external-api.md',
        root / 'references' / 'safety.md',
    ]
    report: dict[str, Any] = {
        'ok': all(item.exists() for item in required),
        'skill': SKILL_NAME,
        'project_id': PROJECT_ID,
        'mode': 'standalone',
        'skill_root': str(root),
        'cwd': str(Path.cwd().resolve()),
        'required_paths': {str(item.relative_to(root)): item.exists() for item in required},
        'api_token_present': bool(os.environ.get('PESQUISA_API_TOKEN')),
        'api_url': os.environ.get('PESQUISA_API_URL', APP_URL),
        'online': None,
    }
    if args.online:
        try:
            request = urllib.request.Request(f'{APP_URL}/api/health', headers={'User-Agent': SKILL_NAME})
            with urllib.request.urlopen(request, timeout=15) as response:
                body = response.read().decode('utf-8', errors='replace')
                report['online'] = {'status': response.status, 'ok': 200 <= response.status < 300, 'body_preview': body[:300]}
        except (urllib.error.URLError, TimeoutError) as exc:
            report['online'] = {'ok': False, 'error': str(exc)}
            report['ok'] = False
    json_print(report)
    return 0 if report['ok'] else 1


def command_capabilities(args: argparse.Namespace) -> int:
    payload = {'skill': SKILL_NAME, 'capabilities': CAPABILITIES}
    if args.json:
        json_print(payload)
    else:
        for domain, actions in CAPABILITIES.items():
            print(domain + ': ' + ', '.join(actions))
    return 0


def command_encoding(args: argparse.Namespace) -> int:
    path = Path(args.file).resolve()
    if not path.exists():
        raise SystemExit(f'Arquivo nao encontrado: {path}')
    issues: list[dict[str, Any]] = []
    values = 0
    try:
        for source, value in iter_file_strings(path, args.sheet):
            values += 1
            issues.extend(text_issues(value, source))
    except UnicodeDecodeError as exc:
        issues.append({'source': str(path), 'kind': 'invalid_utf8', 'detail': str(exc)})
    result = {'ok': not issues, 'file': str(path), 'strings_checked': values, 'issues': issues[:100], 'issue_count': len(issues)}
    json_print(result)
    return 0 if result['ok'] else 2


def read_tabular(path: Path, sheet: str | None = None) -> tuple[str, list[str], list[dict[str, Any]]]:
    if path.suffix.lower() == '.csv':
        with path.open('r', encoding='utf-8-sig', errors='strict', newline='') as handle:
            sample = handle.read(8192)
            handle.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
            except csv.Error:
                dialect = csv.excel
            reader = csv.DictReader(handle, dialect=dialect)
            headers = [str(item or '').strip() for item in (reader.fieldnames or [])]
            rows = [dict(row) for row in reader]
        return 'csv', headers, rows
    if path.suffix.lower() == '.xlsx':
        try:
            import openpyxl
        except ImportError as exc:
            raise SystemExit('Instale openpyxl para auditar XLSX.') from exc
        book = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = book[sheet] if sheet else book.active
        iterator = ws.iter_rows(values_only=True)
        first = next(iterator, ())
        headers = [str(item or '').strip() for item in first]
        rows = [dict(zip(headers, row)) for row in iterator]
        title = ws.title
        book.close()
        return title, headers, rows
    raise SystemExit('sample-audit aceita somente .csv ou .xlsx.')


def choose_column(headers: list[str], explicit: str | None, candidates: tuple[str, ...]) -> str | None:
    lookup = {item.strip().casefold(): item for item in headers}
    if explicit:
        return lookup.get(explicit.strip().casefold())
    for candidate in candidates:
        if candidate.casefold() in lookup:
            return lookup[candidate.casefold()]
    return None


def command_sample_audit(args: argparse.Namespace) -> int:
    path = Path(args.file).resolve()
    if not path.exists():
        raise SystemExit(f'Arquivo nao encontrado: {path}')
    source, headers, rows = read_tabular(path, args.sheet)
    email_col = choose_column(headers, args.email_column, ('email', 'e-mail', 'emailresponsavelfinanceiro', 'email responsavel financeiro', 'email resp fin', 'email institucional', 'email resp acad'))
    community_col = choose_column(headers, args.community_column, ('community_id', 'communityid', 'comunidade'))
    coligada_col = choose_column(headers, None, ('codcoligada', 'cod coligada'))
    filial_col = choose_column(headers, None, ('codfilial', 'cod filial'))
    emails: list[str] = []
    invalid = 0
    empty = 0
    encoding_count = 0
    communities: Counter[str] = Counter()
    pairs: Counter[str] = Counter()
    for index, row in enumerate(rows, 2):
        for header, value in row.items():
            if isinstance(value, str):
                encoding_count += len(text_issues(value, f'{path}:{source}:{index}:{header}'))
        raw_email = str(row.get(email_col, '') or '').strip().lower() if email_col else ''
        if not raw_email:
            empty += 1
        elif EMAIL_RE.match(raw_email):
            emails.append(raw_email)
        else:
            invalid += 1
        if community_col:
            community = str(row.get(community_col, '') or '').strip()
            if community:
                communities[community] += 1
        if coligada_col or filial_col:
            pair = f'{str(row.get(coligada_col, '') or '').strip()}|{str(row.get(filial_col, '') or '').strip()}'
            pairs[pair] += 1
    counts = Counter(emails)
    result = {
        'ok': bool(headers) and email_col is not None and invalid == 0 and encoding_count == 0,
        'file': str(path), 'source': source, 'headers': headers, 'rows': len(rows),
        'columns': {'email': email_col, 'community': community_col, 'codcoligada': coligada_col, 'codfilial': filial_col},
        'emails': {'valid_rows': len(emails), 'unique': len(counts), 'empty': empty, 'invalid': invalid, 'duplicate_rows': sum(value - 1 for value in counts.values() if value > 1)},
        'community_count': len(communities),
        'top_communities': communities.most_common(20),
        'unit_pair_count': len(pairs),
        'top_unit_pairs': pairs.most_common(20),
        'encoding_issue_count': encoding_count,
    }
    json_print(result)
    return 0 if result['ok'] else 2


def operation_risk(operation: str) -> str:
    if operation in EXTERNAL:
        return 'external'
    if operation in DESTRUCTIVE:
        return 'destructive'
    if operation in SCHEMA_OR_DEPLOY:
        return 'schema_or_deploy'
    return 'reversible_write'


def validate_plan_args(value: Any, path: str = '$') -> list[str]:
    problems: list[str] = []
    if isinstance(value, dict):
        for key, item in value.items():
            lowered = str(key).casefold()
            if any(part in lowered for part in FORBIDDEN_KEY_PARTS):
                problems.append(f'{path}.{key}: secret proibido no plano')
            problems.extend(validate_plan_args(item, f'{path}.{key}'))
    elif isinstance(value, list):
        for index, item in enumerate(value):
            problems.extend(validate_plan_args(item, f'{path}[{index}]'))
    elif isinstance(value, str):
        for item in text_issues(value, path):
            problems.append(str(item.get('source')) + ': ' + str(item.get('kind')))
    return problems


def plan_fingerprint(operation: str, arguments: dict[str, Any], environment: str) -> str:
    raw = json.dumps({'version': 1, 'operation': operation, 'arguments': arguments, 'environment': environment}, sort_keys=True, ensure_ascii=False, separators=(',', ':'))
    return hashlib.sha256(raw.encode('utf-8')).hexdigest()


def confirmation_code(risk: str, fingerprint: str) -> str:
    prefix = {'external': 'SEND', 'destructive': 'DELETE', 'schema_or_deploy': 'DEPLOY'}.get(risk, 'APPLY')
    return f'{prefix}-{fingerprint[:12].upper()}'


def command_plan(args: argparse.Namespace) -> int:
    domain, dot, action = args.operation.partition('.')
    if not dot or domain not in CAPABILITIES or action not in CAPABILITIES[domain]:
        raise SystemExit(f'Operacao desconhecida: {args.operation}')
    arguments = json.loads(Path(args.args_file).read_text(encoding='utf-8-sig'))
    if not isinstance(arguments, dict):
        raise SystemExit('args-file deve conter um objeto JSON.')
    problems = validate_plan_args(arguments)
    if problems:
        json_print({'ok': False, 'problems': problems})
        return 2
    risk = operation_risk(args.operation)
    fingerprint = plan_fingerprint(args.operation, arguments, args.environment)
    plan = {
        'version': 1, 'skill': SKILL_NAME, 'project_id': PROJECT_ID,
        'created_at': datetime.now(timezone.utc).isoformat(),
        'environment': args.environment, 'operation': args.operation,
        'risk': risk, 'arguments': arguments, 'fingerprint': fingerprint,
        'confirmation_code': confirmation_code(risk, fingerprint),
        'executed': False,
    }
    target = Path(args.out).resolve()
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(json.dumps(plan, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
    json_print({'ok': True, 'plan': str(target), 'operation': args.operation, 'risk': risk, 'fingerprint': fingerprint, 'confirmation_code': plan['confirmation_code']})
    return 0


def command_verify_plan(args: argparse.Namespace) -> int:
    path = Path(args.plan_file).resolve()
    plan = json.loads(path.read_text(encoding='utf-8-sig'))
    required = {'version', 'skill', 'project_id', 'environment', 'operation', 'risk', 'arguments', 'fingerprint', 'confirmation_code', 'executed'}
    problems = [f'campo ausente: {key}' for key in sorted(required - set(plan))]
    if plan.get('skill') != SKILL_NAME:
        problems.append('skill incorreta')
    if plan.get('project_id') != PROJECT_ID:
        problems.append('project_id incorreto')
    operation = str(plan.get('operation', ''))
    arguments = plan.get('arguments', {})
    environment = str(plan.get('environment', 'production'))
    if isinstance(arguments, dict):
        expected = plan_fingerprint(operation, arguments, environment)
        if plan.get('fingerprint') != expected:
            problems.append('fingerprint invalido: o plano foi alterado')
        risk = operation_risk(operation)
        if plan.get('risk') != risk:
            problems.append('classe de risco invalida')
        if plan.get('confirmation_code') != confirmation_code(risk, expected):
            problems.append('codigo de confirmacao invalido')
        problems.extend(validate_plan_args(arguments))
    else:
        problems.append('arguments deve ser objeto')
    result = {'ok': not problems, 'plan': str(path), 'operation': operation, 'risk': plan.get('risk'), 'confirmation_code': plan.get('confirmation_code'), 'problems': problems}
    json_print(result)
    return 0 if result['ok'] else 2


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(prog='pesquisa-cli', description='Preflight e planos seguros do miniapp Layers.')
    sub = parser.add_subparsers(dest='command', required=True)

    doctor = sub.add_parser('doctor', help='Verificar a skill standalone sem depender do projeto.')
    doctor.add_argument('--online', action='store_true', help='Consultar /api/health em modo somente leitura.')
    doctor.set_defaults(func=command_doctor)

    capabilities = sub.add_parser('capabilities', help='Listar operacoes cobertas.')
    capabilities.add_argument('--json', action='store_true')
    capabilities.set_defaults(func=command_capabilities)

    encoding = sub.add_parser('encoding-check', help='Validar UTF-8, NFC e mojibake.')
    encoding.add_argument('--file', required=True)
    encoding.add_argument('--sheet')
    encoding.set_defaults(func=command_encoding)

    sample = sub.add_parser('sample-audit', help='Auditar XLSX/CSV sem exibir PII.')
    sample.add_argument('--file', required=True)
    sample.add_argument('--sheet')
    sample.add_argument('--email-column')
    sample.add_argument('--community-column')
    sample.set_defaults(func=command_sample_audit)

    plan = sub.add_parser('plan', help='Criar plano imutavel para uma escrita.')
    plan.add_argument('--operation', required=True)
    plan.add_argument('--args-file', required=True)
    plan.add_argument('--out', required=True)
    plan.add_argument('--environment', choices=('production', 'preview', 'local'), default='production')
    plan.set_defaults(func=command_plan)

    verify = sub.add_parser('verify-plan', help='Validar fingerprint e seguranca do plano.')
    verify.add_argument('plan_file')
    verify.set_defaults(func=command_verify_plan)
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    return int(args.func(args))


if __name__ == '__main__':
    sys.exit(main())
